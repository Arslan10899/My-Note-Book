import os
import io
import json
import shutil
import sqlite3
import base64
import uuid
import tempfile
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_from_directory, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import datetime, timedelta
from database import get_db, init_db, seed_data

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'intellinote-dev-secret-2026-do-not-use-in-prod')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'static', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


@app.route('/')
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return render_template('login.html')


@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')


@app.route('/api/dashboard-data')
@login_required
def dashboard_data():
    conn = get_db()
    today = datetime.now().strftime('%Y-%m-%d')

    tasks = conn.execute(
        "SELECT b.*, p.title as page_title FROM blocks b JOIN pages p ON b.page_id = p.id WHERE b.block_type = 'task' ORDER BY b.created_at DESC"
    ).fetchall()

    instructions = conn.execute(
        "SELECT * FROM blocks WHERE block_type = 'instruction' ORDER BY created_at DESC"
    ).fetchall()

    reminders = conn.execute(
        "SELECT b.*, p.title as page_title FROM blocks b JOIN pages p ON b.page_id = p.id WHERE b.block_type = 'reminder' ORDER BY created_at DESC"
    ).fetchall()

    pages = conn.execute("SELECT * FROM pages ORDER BY created_at DESC").fetchall()

    recent_pages = [dict(p) for p in pages]
    for rp in recent_pages:
        block_counts = conn.execute(
            "SELECT block_type, COUNT(*) as cnt FROM blocks WHERE page_id = ? GROUP BY block_type",
            (rp['id'],)
        ).fetchall()
        rp['block_counts'] = {bc['block_type']: bc['cnt'] for bc in block_counts}
        rp['total_blocks'] = sum(bc['cnt'] for bc in block_counts)
        latest_block = conn.execute(
            "SELECT created_at FROM blocks WHERE page_id = ? ORDER BY created_at DESC LIMIT 1",
            (rp['id'],)
        ).fetchone()
        rp['last_modified'] = latest_block['created_at'] if latest_block else None

    conn.close()

    total_tasks = len(tasks)
    pending = sum(1 for t in tasks if (t['task_status'] or 'pending') in ('pending', 'in_progress', 'on_hold'))
    in_progress = sum(1 for t in tasks if t['task_status'] == 'in_progress')
    completed = sum(1 for t in tasks if t['task_status'] == 'completed')
    high_priority = sum(1 for t in tasks if t['task_priority'] == 'high' and (t['task_status'] or 'pending') != 'completed')

    today_tasks = [dict(t) for t in tasks if t['due_date'] == today]
    overdue_tasks = [dict(t) for t in tasks if t['due_date'] and t['due_date'] < today and (t['task_status'] or 'pending') != 'completed']

    upcoming = [dict(r) for r in reminders[:5]]

    return jsonify({
        'stats': {
            'total_tasks': total_tasks,
            'pending': pending,
            'in_progress': in_progress,
            'completed': completed,
            'instructions': len(instructions),
            'pages': len(pages),
            'high_priority': high_priority
        },
        'today_tasks': today_tasks,
        'overdue_tasks': overdue_tasks,
        'upcoming_reminders': upcoming,
        'recent_pages': recent_pages
    })


@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')

    if not username or not password:
        return jsonify({'error': 'Username and password required'}), 400

    conn = get_db()
    user = conn.execute(
        'SELECT * FROM users WHERE username = ?', (username,)
    ).fetchone()
    conn.close()

    if user and check_password_hash(user['password_hash'], password):
        if user['status'] == 'pending':
            return jsonify({'error': 'Your account is pending admin approval. Please wait.'}), 403
        if user['status'] == 'rejected':
            return jsonify({'error': 'Your account request has been declined.'}), 403
        session['user_id'] = user['id']
        session['username'] = user['username']
        return jsonify({
            'success': True,
            'username': user['username'],
            'full_name': user['full_name'] or user['username'],
            'image_path': user['image_path'] or '',
            'role': user['role'] or 'User'
        })

    return jsonify({'error': 'Invalid credentials'}), 401


@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'success': True})


@app.route('/api/register', methods=['POST'])
def api_register():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    full_name = data.get('full_name', '').strip()
    email = data.get('email', '').strip()
    phone = data.get('phone', '').strip()

    if not username or not password or not full_name:
        return jsonify({'error': 'Username, password and full name are required'}), 400

    if len(password) < 4:
        return jsonify({'error': 'Password must be at least 4 characters'}), 400

    conn = get_db()
    existing = conn.execute('SELECT id FROM users WHERE username = ?', (username,)).fetchone()
    if existing:
        conn.close()
        return jsonify({'error': 'Username already taken'}), 400

    conn.execute(
        "INSERT INTO users (username, password_hash, full_name, email, phone, role, status, requested_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (username, generate_password_hash(password), full_name, email, phone, 'User', 'pending', datetime.now().strftime('%Y-%m-%d %H:%M'))
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': 'Registration submitted. Waiting for admin approval.'}), 201


@app.route('/api/admin/pending-users')
@login_required
def admin_pending_users():
    conn = get_db()
    user = conn.execute('SELECT role FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if not user or user['role'] != 'Admin':
        conn.close()
        return jsonify({'error': 'Admin access required'}), 403
    users = conn.execute("SELECT id, username, full_name, email, phone, role, status, requested_at FROM users WHERE status = 'pending' ORDER BY requested_at DESC").fetchall()
    conn.close()
    return jsonify([dict(u) for u in users])


@app.route('/api/admin/all-users')
@login_required
def admin_all_users():
    conn = get_db()
    user = conn.execute('SELECT role FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if not user or user['role'] != 'Admin':
        conn.close()
        return jsonify({'error': 'Admin access required'}), 403
    users = conn.execute("SELECT id, username, full_name, email, phone, role, status, requested_at, image_path FROM users ORDER BY requested_at DESC").fetchall()
    conn.close()
    return jsonify([dict(u) for u in users])


@app.route('/api/admin/approve-user/<int:user_id>', methods=['POST'])
@login_required
def admin_approve_user(user_id):
    conn = get_db()
    admin = conn.execute('SELECT role FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if not admin or admin['role'] != 'Admin':
        conn.close()
        return jsonify({'error': 'Admin access required'}), 403
    data = request.get_json() or {}
    role = data.get('role', 'User')
    conn.execute("UPDATE users SET status = 'approved', role = ? WHERE id = ?", (role, user_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/admin/reject-user/<int:user_id>', methods=['POST'])
@login_required
def admin_reject_user(user_id):
    conn = get_db()
    admin = conn.execute('SELECT role FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if not admin or admin['role'] != 'Admin':
        conn.close()
        return jsonify({'error': 'Admin access required'}), 403
    conn.execute("UPDATE users SET status = 'rejected' WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/admin/delete-user/<int:user_id>', methods=['POST'])
@login_required
def admin_delete_user(user_id):
    conn = get_db()
    admin = conn.execute('SELECT role FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if not admin or admin['role'] != 'Admin':
        conn.close()
        return jsonify({'error': 'Admin access required'}), 403
    if user_id == session['user_id']:
        conn.close()
        return jsonify({'error': 'Cannot delete yourself'}), 400
    conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/user/profile')
@login_required
def get_profile():
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    conn.close()
    if not user:
        return jsonify({'error': 'User not found'}), 404
    return jsonify({
        'id': user['id'],
        'username': user['username'],
        'full_name': user['full_name'] or '',
        'email': user['email'] or '',
        'phone': user['phone'] or '',
        'role': user['role'] or 'Admin',
        'image_path': user['image_path'] or ''
    })


@app.route('/api/user/profile', methods=['PUT'])
@login_required
def update_profile():
    data = request.get_json()
    full_name = data.get('full_name', '')
    email = data.get('email', '')
    phone = data.get('phone', '')
    image_path = data.get('image_path', '')

    conn = get_db()
    conn.execute(
        'UPDATE users SET full_name = ?, email = ?, phone = ?, image_path = ? WHERE id = ?',
        (full_name, email, phone, image_path, session['user_id'])
    )
    conn.commit()
    user = conn.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    conn.close()
    return jsonify({
        'success': True,
        'full_name': user['full_name'],
        'email': user['email'],
        'phone': user['phone'],
        'image_path': user['image_path']
    })


@app.route('/api/user/change-password', methods=['POST'])
@login_required
def change_password():
    data = request.get_json()
    current_password = data.get('current_password', '')
    new_password = data.get('new_password', '')

    if not current_password or not new_password:
        return jsonify({'error': 'Both current and new password required'}), 400

    if len(new_password) < 4:
        return jsonify({'error': 'Password must be at least 4 characters'}), 400

    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()

    if not user or not check_password_hash(user['password_hash'], current_password):
        conn.close()
        return jsonify({'error': 'Current password is incorrect'}), 401

    conn.execute(
        'UPDATE users SET password_hash = ? WHERE id = ?',
        (generate_password_hash(new_password), session['user_id'])
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': 'Password updated successfully'})


@app.route('/api/pages', methods=['GET'])
@login_required
def get_pages():
    conn = get_db()
    pages = conn.execute(
        'SELECT * FROM pages ORDER BY created_at DESC'
    ).fetchall()
    conn.close()
    return jsonify([dict(page) for page in pages])


@app.route('/api/pages', methods=['POST'])
@login_required
def create_page():
    data = request.get_json()
    title = data.get('title', 'Untitled')
    icon = data.get('icon', '📄')

    conn = get_db()
    cursor = conn.execute(
        'INSERT INTO pages (title, icon) VALUES (?, ?)',
        (title, icon)
    )
    page_id = cursor.lastrowid
    conn.commit()
    page = conn.execute('SELECT * FROM pages WHERE id = ?', (page_id,)).fetchone()
    conn.close()
    return jsonify(dict(page)), 201


@app.route('/api/pages/<int:page_id>', methods=['GET'])
@login_required
def get_page(page_id):
    conn = get_db()
    page = conn.execute('SELECT * FROM pages WHERE id = ?', (page_id,)).fetchone()
    if not page:
        conn.close()
        return jsonify({'error': 'Page not found'}), 404

    blocks = conn.execute(
        'SELECT * FROM blocks WHERE page_id = ? ORDER BY created_at',
        (page_id,)
    ).fetchall()
    conn.close()

    result = dict(page)
    result['blocks'] = [dict(block) for block in blocks]
    return jsonify(result)


@app.route('/api/pages/<int:page_id>', methods=['PUT'])
@login_required
def update_page(page_id):
    data = request.get_json()
    conn = get_db()
    page = conn.execute('SELECT * FROM pages WHERE id = ?', (page_id,)).fetchone()
    if not page:
        conn.close()
        return jsonify({'error': 'Page not found'}), 404

    title = data.get('title', page['title'])
    icon = data.get('icon', page['icon'])

    conn.execute(
        'UPDATE pages SET title = ?, icon = ? WHERE id = ?',
        (title, icon, page_id)
    )
    conn.commit()
    updated = conn.execute('SELECT * FROM pages WHERE id = ?', (page_id,)).fetchone()
    conn.close()
    return jsonify(dict(updated))


@app.route('/api/pages/<int:page_id>', methods=['DELETE'])
@login_required
def delete_page(page_id):
    conn = get_db()
    conn.execute('DELETE FROM pages WHERE id = ?', (page_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/upload-image', methods=['POST'])
@login_required
def upload_image():
    if 'image' not in request.files:
        return jsonify({'error': 'No image file'}), 400

    file = request.files['image']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp'):
        return jsonify({'error': 'Invalid image format'}), 400

    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    return jsonify({'success': True, 'image_path': f'/static/uploads/{filename}'}), 201


@app.route('/api/upload-image-base64', methods=['POST'])
@login_required
def upload_image_base64():
    data = request.get_json()
    image_data = data.get('image_data', '')

    if not image_data:
        return jsonify({'error': 'No image data'}), 400

    if ',' in image_data:
        image_data = image_data.split(',', 1)[1]

    ext = 'png'
    if 'jpeg' in image_data[:20] or 'jpg' in image_data[:20]:
        ext = 'jpg'

    try:
        image_bytes = base64.b64decode(image_data)
    except Exception:
        return jsonify({'error': 'Invalid base64 data'}), 400

    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    with open(filepath, 'wb') as f:
        f.write(image_bytes)

    return jsonify({'success': True, 'image_path': f'/static/uploads/{filename}'}), 201


@app.route('/api/blocks', methods=['POST'])
@login_required
def create_block():
    data = request.get_json()
    page_id = data.get('page_id')
    block_type = data.get('block_type')
    content = data.get('content', '')
    subject = data.get('subject', '')
    due_date = data.get('due_date')
    recurrence = data.get('recurrence', 'none')
    ref_date = data.get('ref_date')
    ref_type = data.get('ref_type', 'none')
    ref_detail = data.get('ref_detail', '')
    priority = data.get('priority', 'normal')
    is_active = data.get('is_active', 1)
    end_date = data.get('end_date')
    image_path = data.get('image_path', '')
    task_priority = data.get('task_priority', 'medium')
    category = data.get('category', 'general')
    description = data.get('description', '')
    task_status = data.get('task_status', 'pending')
    reminder_time = data.get('reminder_time', '')

    if not page_id or not block_type:
        return jsonify({'error': 'page_id and block_type required'}), 400

    conn = get_db()
    cursor = conn.execute(
        '''INSERT INTO blocks (page_id, block_type, content, subject, due_date, recurrence, ref_date, ref_type, ref_detail, priority, is_active, end_date, image_path, task_priority, category, description, task_status, reminder_time)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
        (page_id, block_type, content, subject, due_date, recurrence, ref_date, ref_type, ref_detail, priority, is_active, end_date, image_path, task_priority, category, description, task_status, reminder_time)
    )
    block_id = cursor.lastrowid
    conn.commit()
    block = conn.execute('SELECT * FROM blocks WHERE id = ?', (block_id,)).fetchone()
    conn.close()
    return jsonify(dict(block)), 201


@app.route('/api/blocks/<int:block_id>', methods=['PUT'])
@login_required
def update_block(block_id):
    data = request.get_json()
    conn = get_db()
    block = conn.execute('SELECT * FROM blocks WHERE id = ?', (block_id,)).fetchone()
    if not block:
        conn.close()
        return jsonify({'error': 'Block not found'}), 404

    content = data.get('content', block['content'])
    subject = data.get('subject', block['subject'])
    status = data.get('status', block['status'])
    due_date = data.get('due_date', block['due_date'])
    recurrence = data.get('recurrence', block['recurrence'])
    ref_date = data.get('ref_date', block['ref_date'])
    ref_type = data.get('ref_type', block['ref_type'])
    ref_detail = data.get('ref_detail', block['ref_detail'])
    priority = data.get('priority', block['priority'])
    is_active = data.get('is_active', block['is_active'])
    end_date = data.get('end_date', block['end_date'])
    image_path = data.get('image_path', block['image_path'])
    task_priority = data.get('task_priority', block['task_priority'])
    category = data.get('category', block['category'])
    description = data.get('description', block['description'])
    task_status = data.get('task_status', block['task_status'])
    reminder_time = data.get('reminder_time', block['reminder_time'] if 'reminder_time' in block.keys() else '')

    conn.execute(
        '''UPDATE blocks SET content = ?, subject = ?, status = ?, due_date = ?, recurrence = ?,
        ref_date = ?, ref_type = ?, ref_detail = ?, priority = ?, is_active = ?, end_date = ?, image_path = ?,
        task_priority = ?, category = ?, description = ?, task_status = ?, reminder_time = ? WHERE id = ?''',
        (content, subject, status, due_date, recurrence, ref_date, ref_type, ref_detail, priority, is_active, end_date, image_path, task_priority, category, description, task_status, reminder_time, block_id)
    )
    conn.commit()
    updated = conn.execute('SELECT * FROM blocks WHERE id = ?', (block_id,)).fetchone()
    conn.close()
    return jsonify(dict(updated))


@app.route('/api/blocks/<int:block_id>', methods=['DELETE'])
@login_required
def delete_block(block_id):
    conn = get_db()
    conn.execute('DELETE FROM blocks WHERE id = ?', (block_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/search')
@login_required
def search():
    query = request.args.get('q', '')
    if not query:
        return jsonify([])

    conn = get_db()
    search_term = f'%{query}%'

    pages = conn.execute(
        "SELECT id, title, icon, 'page' as result_type FROM pages WHERE title LIKE ?",
        (search_term,)
    ).fetchall()

    instruction_blocks = conn.execute(
        """SELECT b.id, b.content, b.subject, b.ref_type, b.ref_detail, b.priority, b.is_active, p.title as page_title, 'instruction' as result_type
           FROM blocks b JOIN pages p ON b.page_id = p.id
           WHERE b.block_type = 'instruction' AND (b.content LIKE ? OR b.subject LIKE ? OR b.ref_detail LIKE ?)""",
        (search_term, search_term, search_term)
    ).fetchall()

    tasks = conn.execute(
        """SELECT b.id, b.content, b.status, b.due_date, p.title as page_title, 'task' as result_type
           FROM blocks b JOIN pages p ON b.page_id = p.id
           WHERE b.block_type = 'task' AND b.content LIKE ?""",
        (search_term,)
    ).fetchall()

    reminders = conn.execute(
        """SELECT b.id, b.content, b.recurrence, p.title as page_title, 'reminder' as result_type
           FROM blocks b JOIN pages p ON b.page_id = p.id
           WHERE b.block_type = 'reminder' AND b.content LIKE ?""",
        (search_term,)
    ).fetchall()

    results = [dict(r) for r in pages] + [dict(r) for r in instruction_blocks] + [dict(r) for r in tasks] + [dict(r) for r in reminders]
    conn.close()
    return jsonify(results)


@app.route('/api/search/all')
@login_required
def search_all():
    conn = get_db()

    pages = conn.execute("SELECT id, title, icon FROM pages ORDER BY created_at DESC").fetchall()
    blocks = conn.execute(
        """SELECT b.id, b.page_id, b.block_type, b.content, b.subject, b.status, b.due_date,
           b.recurrence, b.ref_type, b.ref_detail, b.priority, b.is_active, b.end_date, b.image_path,
           p.title as page_title, p.icon as page_icon
           FROM blocks b JOIN pages p ON b.page_id = p.id ORDER BY b.created_at DESC"""
    ).fetchall()
    conn.close()

    return jsonify({
        'pages': [dict(p) for p in pages],
        'blocks': [dict(b) for b in blocks]
    })


@app.route('/api/assistant/briefing')
@login_required
def assistant_briefing():
    today = datetime.now().date()
    yesterday = today - timedelta(days=1)
    briefing = []

    conn = get_db()

    daily_tasks = conn.execute(
        """SELECT b.content, b.status, b.due_date
           FROM blocks b
           WHERE b.block_type = 'task' AND b.due_date = ?""",
        (today.strftime('%Y-%m-%d'),)
    ).fetchall()

    for task in daily_tasks:
        briefing.append({
            'type': 'daily_task',
            'content': task['content'],
            'status': task['status'],
            'message': f"📋 Today's Task: {task['content']}"
        })

    missed_attendance = conn.execute(
        """SELECT status FROM logs
           WHERE log_type = 'attendance' AND log_date = ?""",
        (yesterday.strftime('%Y-%m-%d'),)
    ).fetchone()

    if missed_attendance and missed_attendance['status'] == 'missed':
        briefing.append({
            'type': 'missed_alert',
            'content': 'Submit daily attendance',
            'message': "⚠️ Yesterday's attendance was not recorded. Please submit 2 days' attendance today."
        })

    if today.weekday() == 0:
        weekly_tasks = conn.execute(
            """SELECT content FROM blocks
               WHERE block_type = 'reminder' AND recurrence = 'weekly_monday'"""
        ).fetchall()
        for task in weekly_tasks:
            briefing.append({
                'type': 'weekly_alert',
                'content': task['content'],
                'message': f"📅 Weekly Reminder: {task['content']}"
            })

    daily_reminders = conn.execute(
        """SELECT content FROM blocks
           WHERE block_type = 'reminder' AND recurrence = 'daily'"""
    ).fetchall()
    for reminder in daily_reminders:
        briefing.append({
            'type': 'daily_reminder',
            'content': reminder['content'],
            'message': f"🔔 Daily Reminder: {reminder['content']}"
        })

    conn.close()
    return jsonify(briefing)


@app.route('/api/logs/attendance', methods=['POST'])
@login_required
def log_attendance():
    data = request.get_json()
    date = data.get('date', datetime.now().strftime('%Y-%m-%d'))
    status = data.get('status', 'submitted')

    conn = get_db()
    existing = conn.execute(
        "SELECT id FROM logs WHERE log_type = 'attendance' AND log_date = ?",
        (date,)
    ).fetchone()

    if existing:
        conn.execute(
            "UPDATE logs SET status = ? WHERE id = ?",
            (status, existing['id'])
        )
    else:
        conn.execute(
            "INSERT INTO logs (log_type, log_date, status) VALUES (?, ?, ?)",
            ('attendance', date, status)
        )

    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/admin/backup-stats')
@login_required
def admin_backup_stats():
    conn = get_db()
    admin = conn.execute('SELECT role FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if not admin or admin['role'] != 'Admin':
        conn.close()
        return jsonify({'error': 'Admin access required'}), 403
    pages = conn.execute('SELECT COUNT(*) as c FROM pages').fetchone()['c']
    blocks = conn.execute('SELECT COUNT(*) as c FROM blocks').fetchone()['c']
    users = conn.execute('SELECT COUNT(*) as c FROM users').fetchone()['c']
    logs_count = conn.execute('SELECT COUNT(*) as c FROM logs').fetchone()['c']
    conn.close()
    return jsonify({'pages': pages, 'blocks': blocks, 'users': users, 'logs': logs_count})


@app.route('/api/admin/backup/json', methods=['GET'])
@login_required
def admin_backup_json():
    conn = get_db()
    admin = conn.execute('SELECT role FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if not admin or admin['role'] != 'Admin':
        conn.close()
        return jsonify({'error': 'Admin access required'}), 403

    pages = [dict(p) for p in conn.execute('SELECT * FROM pages ORDER BY id').fetchall()]
    blocks = [dict(b) for b in conn.execute('SELECT * FROM blocks ORDER BY id').fetchall()]
    users = [dict(u) for u in conn.execute('SELECT id, username, full_name, email, phone, role, image_path, status, requested_at FROM users ORDER BY id').fetchall()]
    logs = [dict(l) for l in conn.execute('SELECT * FROM logs ORDER BY id').fetchall()]
    conn.close()

    backup = {
        'version': '1.0', 'app': 'IntelliNote',
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'created_by': session.get('username', 'admin'),
        'data': {'pages': pages, 'blocks': blocks, 'users': users, 'logs': logs}
    }
    filename = f"intellinote-backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}.json"
    return send_file(io.BytesIO(json.dumps(backup, indent=2, default=str).encode('utf-8')), mimetype='application/json', as_attachment=True, download_name=filename)


@app.route('/api/admin/backup/sqlite', methods=['GET'])
@login_required
def admin_backup_sqlite():
    conn = get_db()
    admin = conn.execute('SELECT role FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if not admin or admin['role'] != 'Admin':
        conn.close()
        return jsonify({'error': 'Admin access required'}), 403
    conn.close()

    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'intellinote.db')
    filename = f"intellinote-backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}.db"
    return send_file(db_path, mimetype='application/x-sqlite3', as_attachment=True, download_name=filename)


@app.route('/api/admin/backup/excel', methods=['GET'])
@login_required
def admin_backup_excel():
    conn = get_db()
    admin = conn.execute('SELECT role FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if not admin or admin['role'] != 'Admin':
        conn.close()
        return jsonify({'error': 'Admin access required'}), 403

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='5856D6', end_color='5856D6', fill_type='solid')
    border = Border(bottom=Side(style='thin', color='CCCCCC'))
    align = Alignment(wrap_text=True, vertical='top')

    pages = conn.execute('SELECT * FROM pages ORDER BY id').fetchall()
    ws = wb.active
    ws.title = 'Pages'
    page_headers = ['ID', 'Title', 'Icon', 'Created At']
    ws.append(page_headers)
    for col_idx, h in enumerate(page_headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
    for p in pages:
        ws.append([p['id'], p['title'], p['icon'], p['created_at']])
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['D'].width = 20

    blocks = conn.execute('SELECT * FROM blocks ORDER BY id').fetchall()
    ws2 = wb.create_sheet('Blocks')
    block_headers = ['ID', 'Page ID', 'Type', 'Content', 'Subject', 'Status', 'Due Date', 'Recurrence',
                     'Ref Date', 'Ref Type', 'Ref Detail', 'Priority', 'Active', 'End Date', 'Image Path',
                     'Task Priority', 'Category', 'Description', 'Task Status', 'Reminder Time', 'Created At']
    ws2.append(block_headers)
    for col_idx, h in enumerate(block_headers, 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
    for b in blocks:
        ws2.append([b['id'], b['page_id'], b['block_type'], b['content'], b['subject'], b['status'],
                     b['due_date'], b['recurrence'], b['ref_date'], b['ref_type'], b['ref_detail'],
                     b['priority'], b['is_active'], b['end_date'], b['image_path'], b['task_priority'],
                     b['category'], b['description'], b['task_status'], b['reminder_time'], b['created_at']])
    ws2.column_dimensions['D'].width = 60
    ws2.column_dimensions['E'].width = 30

    users = conn.execute('SELECT id, username, full_name, email, phone, role, image_path, status, requested_at FROM users ORDER BY id').fetchall()
    ws3 = wb.create_sheet('Users')
    user_headers = ['ID', 'Username', 'Full Name', 'Email', 'Phone', 'Role', 'Image Path', 'Status', 'Requested At']
    ws3.append(user_headers)
    for col_idx, h in enumerate(user_headers, 1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
    for u in users:
        ws3.append([u['id'], u['username'], u['full_name'], u['email'], u['phone'], u['role'], u['image_path'], u['status'], u['requested_at']])
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 25

    logs = conn.execute('SELECT * FROM logs ORDER BY id').fetchall()
    ws4 = wb.create_sheet('Logs')
    log_headers = ['ID', 'Log Type', 'Date', 'Status']
    ws4.append(log_headers)
    for col_idx, h in enumerate(log_headers, 1):
        cell = ws4.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
    for l in logs:
        ws4.append([l['id'], l['log_type'], l['log_date'], l['status']])

    conn.close()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"intellinote-backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)


@app.route('/api/admin/restore/json', methods=['POST'])
@login_required
def admin_restore_json():
    conn = get_db()
    admin = conn.execute('SELECT role FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if not admin or admin['role'] != 'Admin':
        conn.close()
        return jsonify({'error': 'Admin access required'}), 403

    if 'backup_file' not in request.files:
        conn.close()
        return jsonify({'error': 'No backup file provided'}), 400

    file = request.files['backup_file']
    try:
        backup = json.loads(file.read().decode('utf-8'))
    except Exception:
        conn.close()
        return jsonify({'error': 'Invalid JSON backup file'}), 400

    if backup.get('app') != 'IntelliNote':
        conn.close()
        return jsonify({'error': 'Not a valid IntelliNote backup file'}), 400

    mode = request.form.get('mode', 'merge')

    try:
        cursor = conn.cursor()
        if mode == 'replace':
            cursor.execute('DELETE FROM blocks')
            cursor.execute('DELETE FROM pages')
            cursor.execute('DELETE FROM logs')

        data = backup.get('data', {})
        old_to_new_page = {}
        for page in data.get('pages', []):
            if mode == 'replace':
                cursor.execute('INSERT INTO pages (id, title, icon, created_at) VALUES (?, ?, ?, ?)',
                    (page['id'], page['title'], page.get('icon', '📄'), page.get('created_at', '')))
                old_to_new_page[page['id']] = cursor.lastrowid
            else:
                existing = cursor.execute('SELECT id FROM pages WHERE id = ?', (page['id'],)).fetchone()
                if existing:
                    cursor.execute('UPDATE pages SET title = ?, icon = ? WHERE id = ?',
                        (page['title'], page.get('icon', '📄'), page['id']))
                    old_to_new_page[page['id']] = page['id']
                else:
                    cursor.execute('INSERT INTO pages (id, title, icon, created_at) VALUES (?, ?, ?, ?)',
                        (page['id'], page['title'], page.get('icon', '📄'), page.get('created_at', '')))
                    old_to_new_page[page['id']] = cursor.lastrowid

        for block in data.get('blocks', []):
            new_pid = old_to_new_page.get(block['page_id'], block['page_id'])
            cursor.execute('''INSERT INTO blocks (page_id, block_type, content, subject, status, due_date,
                recurrence, ref_date, ref_type, ref_detail, priority, is_active, end_date,
                image_path, task_priority, category, description, task_status, reminder_time, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (new_pid, block.get('block_type'), block.get('content', ''), block.get('subject', ''),
                 block.get('status', 'pending'), block.get('due_date'), block.get('recurrence'),
                 block.get('ref_date'), block.get('ref_type', 'none'), block.get('ref_detail', ''),
                 block.get('priority', 'normal'), block.get('is_active', 1), block.get('end_date'),
                 block.get('image_path', ''), block.get('task_priority', 'medium'), block.get('category', 'general'),
                 block.get('description', ''), block.get('task_status', 'pending'),
                 block.get('reminder_time', ''), block.get('created_at', '')))

        for log in data.get('logs', []):
            cursor.execute('INSERT INTO logs (log_type, log_date, status) VALUES (?, ?, ?)',
                (log.get('log_type'), log.get('log_date'), log.get('status')))

        conn.commit()
        conn.close()
        stats = backup.get('stats', {})
        return jsonify({'success': True, 'message': f'Restored ({mode} mode)',
            'restored': {'pages': stats.get('total_pages', len(data.get('pages', []))),
                         'blocks': stats.get('total_blocks', len(data.get('blocks', []))),
                         'logs': stats.get('total_logs', len(data.get('logs', [])))}})
    except Exception as e:
        conn.close()
        return jsonify({'error': f'Restore failed: {str(e)}'}), 500


@app.route('/api/admin/restore/sqlite', methods=['POST'])
@login_required
def admin_restore_sqlite():
    conn = get_db()
    admin = conn.execute('SELECT role FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if not admin or admin['role'] != 'Admin':
        conn.close()
        return jsonify({'error': 'Admin access required'}), 403
    conn.close()

    if 'backup_file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['backup_file']
    if not file.filename.endswith('.db'):
        return jsonify({'error': 'File must be a .db SQLite file'}), 400

    content = file.read()
    if len(content) < 16:
        return jsonify({'error': 'File too small to be a valid SQLite database'}), 400

    tmp_path = None
    try:
        tmp_fd, tmp_path = tempfile.mkstemp(suffix='.db')
        os.close(tmp_fd)
        with open(tmp_path, 'wb') as f:
            f.write(content)
        test_conn = sqlite3.connect(tmp_path)
        test_tables = [r[0] for r in test_conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        test_conn.close()
        if 'pages' not in test_tables or 'blocks' not in test_tables:
            return jsonify({'error': 'Invalid database: missing pages/blocks tables'}), 400
    except Exception:
        return jsonify({'error': 'Invalid SQLite database file'}), 400
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)

    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'intellinote.db')
    backup_path = db_path + '.bak'
    shutil.copy2(db_path, backup_path)
    try:
        with open(db_path, 'wb') as f:
            f.write(content)
        init_db()
        if os.path.exists(backup_path):
            os.remove(backup_path)
        return jsonify({'success': True, 'message': 'Database replaced successfully. Please refresh.', 'requires_refresh': True})
    except Exception as e:
        shutil.copy2(backup_path, db_path)
        return jsonify({'error': f'Restore failed: {str(e)}. Original restored.'}), 500


if __name__ == '__main__':
    init_db()
    seed_data()
    app.run(debug=False, port=5000)
